import openpyxl
def import_files(startPoint, filename):
    '''
    Import all the files
    
    Parameters:
    roi = 'data/roiNumbering.txt'
    minima = 'data/localMinima.txt'
    analyze_particle = 'data/aParticle.txt'
    '''
    roi_T = f'data/roi_numbering/{filename}_T.txt'
    roi_A = f'data/roi_numbering/{filename}_A.txt'
    analyze_particle_T = f'data/analyze_particle/{filename}_T.txt'
    analyze_particle_A = f'data/analyze_particle/{filename}_A.txt'
    
    #Read roiNumbering.txt
    try:
        with open(roi_T, 'r') as file:
            xT = []
            for line in file:
                values = line.strip().split('\t')
                xT.append(int(values[0]))
        
        with open(roi_A, 'r') as file:
            xA = []
            for line in file:
                values = line.strip().split('\t')
                xA.append(int(values[0]))

        # Read the result of analyze particle; aParticle.txt
        with open(analyze_particle_T, 'r') as file:
            # Read the first line (the header) and split it on the tab character
            header = file.readline().strip().split('\t')
            # initialize empty lists to store the data
            center_T_x = []
            center_T_y = []
            # iterate over each line in the file
            for line in file:
                # split the line by the tab character
                values = line.strip().split('\t')
                # add the values to the appropriate list
                center_T_x.append((float(values[5])))
                center_T_y.append((float(values[6])))

        with open(analyze_particle_A, 'r') as file:
            # Read the first line (the header) and split it on the tab character
            header = file.readline().strip().split('\t')
            # initialize empty lists to store the data
            center_A_x = []
            center_A_y = []
            # iterate over each line in the file
            for line in file:
                # split the line by the tab character
                values = line.strip().split('\t')
                # add the values to the appropriate list
                center_A_x.append((float(values[5])))
                center_A_y.append((float(values[6])))
    

        return (xA, xT, center_T_x, center_T_y, center_A_x, center_A_y)

    except FileNotFoundError:
        print("File not found. Check again.")
        exit()    

def import_boundaries(startPoint, filename):
    local_minima = f'data/localminima/{filename}.txt'
    with open(local_minima, 'r') as file:
            # initialize empty lists to store the data
            left = []
            right = []
            for line in file:
                    values = line.strip().split('\t')
                    left.append(int(values[0])+startPoint)
                    if filename[-1] == "2":
                        right.append(int(values[1])+startPoint)
    if right != []:
        return left, right
    else:
        return left

def empty_list(length):
    '''Creates empty list based on the length of the data'''
    return ['']*length

def average(lst):
    return sum(lst)/len(lst)

from bisect import bisect_left

def take_closest(myList, myNumber):
    """
    Assumes myList is sorted. Returns closest value to myNumber.

    If two numbers are equally close, return the smallest number.
    """
    pos = bisect_left(myList, myNumber)
    if pos == 0:
        return myList[0]
    if pos == len(myList):
        return myList[-1]
    before = myList[pos - 1]
    after = myList[pos]
    if after - myNumber < myNumber - before:
        return after
    else:
        return before

def data_with_empty_list(data, source):
    '''
    Add the data into destination if its in the source
    data = data_collision
    source = fitted_minima
    destination = collision
    '''
    temp = source.copy()

    for i, num in enumerate(temp):
        if num not in data:
            temp[i] = ''
    
    return temp


# Not implemented anymore
def save_collision(x, minima, nonCollision, collision, initial, ndata, filename):
    result = []
    for i in range(0, ndata):
        result.append([])
        result[i].append(x[i])
        result[i].append(minima[i])
        result[i].append(nonCollision[i])
        result[i].append(collision[i])
        if i < len(initial):
            result[i].append(initial[i])
        else:
            result[i].append('')

    # Saves the result into a .txt file
    with open(f"data/collision_result/{filename}.txt", "w") as file:
        for row in result:
            file.write('\t'.join(map(str,row))+'\n')

    print(f"Data saved to data/collision_result/{filename}.txt")

def save_for_graph(x, minima, collision, initial, ndata, filename):
    '''Save the results to make graph on excel'''
    result = []
    for i in range(0, ndata):
        result.append([])
        result[i].append(x[i])
        result[i].append(minima[i])
        result[i].append(collision[i])
        if i < len(initial):
            result[i].append(initial[i])
        else:
            result[i].append('')

    # Saves the result into a .txt file
    with open(f"data/collision_result/{filename}.txt", "w") as file:
        for row in result:
            file.write('\t'.join(map(str,row))+'\n')

    print(f"Data saved to data/collision_result/{filename}.txt")

def convert_data(data, source):    
    '''
    Converts the data comprised of index into the actual grey value.

    Parameters:
    data (list): The input data to be processed, usually 'collision'
    source (list): The reference data comprised of grey value

    Returns:
    temp(list): The converted list with grey value
    '''
    temp = []
    for item in data:
        if item != '':
            temp.append(source[item-1])
        else:
            temp.append('')
    return temp

def slice_data(data, percentage):
    """
    Group and slice the input data.
    
    Parameters:
    data (list): The input data to be processed
    threshold (int): The threshold value to group the data
    percentage (float): The percentage of each group to retain
    
    Returns:
    list: The processed data
    """
    
    new_data =[]
    for group in data:
        slice_count = int(percentage * len(group) / 100)
        middle = int(len(group)/2)
        front_slice = middle - int(slice_count/2)
        back_slice = middle + int(slice_count/2)

        new_data.extend(group[front_slice:back_slice+1])
    
    return new_data        

def group_data(data):
    """ Group data that has values close to each other"""
    result = []
    for group in data:
        for item in group:
            result.append[item]
    return result


def find_initial_state(minima):
    '''Find the left boundary position prior to droplet movement'''
    normalized_minima = normalize_data(minima)

    ndata = len(minima)
    temp = []
    result = []
    for index in range(ndata):
        temp.append(normalized_minima[index])
        avg = average(temp)
        if normalized_minima[index+5] >= 0.7 * avg:
            result.append(index)
        else:
            break
    return result

def average_5_adj(data):
    ''' Find the average value of a cell with 5 adjacent values'''
    normalized_data = normalize_data(data)
    length_data = len(data)
    average_5_adjacent = []
    temp = 0
    percentage_change = []
    for i in range(0,length_data-1):
        if normalized_data[i] == 0:
            percentage_change.append(0)
        else:
            temp = ((normalized_data[i+1]-normalized_data[i])/normalized_data[i])*100
            #print(i,temp)
            percentage_change.append(temp)

    for i in range(5,length_data-5):
        average_5_adjacent.append(average(percentage_change[i-5:i+6]))
    return average_5_adjacent

def find_extreme_change(data, start, threshold, x):
    '''
    Returns index of X before extreme change
    threshold = threshold percentage as an indicator
    
    Parameters:
    Data: Average of a data along with 5 data points adjacent to it.
    Start: The first point of iteration
    Threshold: Value for extreme change detection

    Example:  1 1 1 1 1 2 5 5 5 5 5
    Return: 32/11
    '''
    length_data = len(data)
    count = 0
    result = 0
    for i in range(start, length_data):
        if abs(data[i]) >= threshold:
            count +=1
        else:
            count = 0
        
        if count == 5:
            result = i - count + 5
            print("Detected extreme change:", x[result])
            break
    return result

def find_data_collision(data, start_point, threshold, x):
    '''
    Find the data collision from percentage change
    '''
    wave_period = 360
    x_start_point = x[start_point] + wave_period // 4
    start_point = x.index(take_closest(x, x_start_point)) - 5
    collision_range = []
    result = []
    n = 0
    end_point = 1

    while end_point != 0:
        end_point = find_extreme_change(data, start_point, threshold, x)
        
        x_end_point = x[end_point]
        if end_point == 0:
            x_end_point = x[-1]

        print(f"Detected area: {x_start_point}-{x_end_point}")
        collision_range.append([x_start_point,x_end_point])
        result.append([])
        for i in range(x_start_point, x_end_point):
            if i in x:
                result[n].append(i)
                
        x_start_point = x_end_point + wave_period // 2
        
        closest_start_point = take_closest(x, x_start_point)
        
        start_point = x.index(closest_start_point) - 5

        n += 1

    return result, collision_range

def normalize_data(data):
    '''Normalize data'''
    minimum_data = min(data)
    temp = []
    for item in data:
        temp.append(item-minimum_data)
    
    return temp

import openpyxl

def save_data_to_excel(data):
    try:
        # Read existing excel file
        wb = openpyxl.load_workbook("result.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        # Create a new excel file if it doesn't exist
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["filename", "fusion_check", "n_collision", "LR length","mean_radius_is", "mean_collision", "mean_is"])

    # Find the first empty row
    next_row = sheet.max_row + 1

    # Append new data to the next empty row
    for column, value in enumerate(data, 1):
        sheet.cell(row=next_row, column=column, value=value)

    # Save workbook
    wb.save("result.xlsx")

import numpy as np
import pyperclip
# makeRectangle(266, 192, 139, 106);
# makeRectangle(x, y, width, height)
def fit_plot(x, y, rct):
    '''
    Fit a straight line based on the center of the trapped droplet
    
    Parameters:
    filename: str
    rct: list
    '''
    width = rct[2]
    x = np.array(x)
    y = np.array(y)
    # calculate polynomial 
    z = np.polyfit(x, y, 1) 
    f = np.poly1d(z) 
    # calculate new x's and y's 
    x_new = np.linspace(x[0], x[-1], 50) 
    y_new = f(x_new)  
    b = round(z[0],2)
    a = round(z[1],2)
    # Equation of fitted line 
    print("Equation of fitted line: y=", b,"x +", a)
    # Parameters for makeLine
    x1 = float(rct[0])
    y1 =  x1*b + a
    x2 = x1 + width
    y2 = x2*b + a
    line_equation = f'makeLine({x1},{y1},{x2},{y2});'
    print(line_equation)
    pyperclip.copy(line_equation)
    return ([x1, y1, x2, y2])

def roi_numbering(filename):
    '''Get the frames detected by ImageJ Analyze Particle function'''
    roi = f'data/roi_file/{filename}.txt'
    with open(roi, 'r') as file:
        header = file.readline().strip().split('\t')
        # initialize empty lists to store the data
        data = []
        # iterate over each line in the file
        for line in file:
            # split the line by the tab character
            values = line.strip().split('\t')
            temp = values[1]
            data.append(int(temp[:4]))
    
    with open(f'data/roi_numbering/{filename}.txt', "w") as file:
        for item in data:
            file.write(str(item)+"\n")

import itertools
from matplotlib import pyplot as plt

def get_local_minima(filename):
    '''Import grey value of collision data'''
    local_minima = f'data/lineplot/{filename}.txt'
    with open(local_minima, 'r') as file:
        # Read the first line (the header) and split it on the tab character
        header = file.readline().strip().split('\t')
        # Create empty lists to store the x and y values
        x = []
        y = []
        nColumn = len(header)
        for i in range(nColumn):
            y.append([])
        # Iterate over the remaining lines in the file
        for row in file:
            # Split the line on the tab character and store the resulting values in the x and y lists
            values = row.strip().split('\t')
            x.append(float(values[0]))
            for i in range(0, nColumn):
                y[i].append(float(values[i+1]))
        result = []

        for i in range(nColumn):
            temp = find_left_right_point(x,y[i])
            if filename[-1] == "1":
                a, b = temp
                result.append([a])
            else:
                a, b = temp
                result.append([a, b])
        
        with open(f"data/localminima/{filename}.txt", "w") as f:
            for item in result:
                temp = "\t".join(map(str,list(itertools.chain(*item))))
                f.write(str(temp) + "\n")
        return 

def find_left_right_point(x,y):
    # Does a single calculation to find the local minima of the data
    y = np.array(y)
    n_data = len(x)
    middle = n_data//2

    # Find the minimum point at the left side
    left = []
    temp = 0
    for i in range(middle):
        if i == 0:
            temp = i
        elif y[i] < y[temp]:
            temp = i

    left.append(temp)

    #Find the maximum point at the right side
    right = []
    temp = middle
    for i in range(middle, n_data):
        if i == middle:
            temp = i
        elif y[i] < y[temp]:
            temp = i
    
    right.append(temp)

    return(left, right)

def plot_minima(x,y):
    # Plot a single data
    result = find_left_right_point(y)
    y = np.array(y)
    minima = result
    print(minima)
    # Plotting the graph
    plt.plot(x,y)
    plt.plot(x[minima[0]],y[minima[0]], "o")
    plt.show()
    
def check_plot(y,i):
    plot_minima(y[i-1])
    # Plot the graph so that its window only opens for 6 seconds
    plt.show()


def plot_everything(nColumn):
    # Plot every available data entry
    for i in range(600,nColumn,1):
        check_plot(i)

